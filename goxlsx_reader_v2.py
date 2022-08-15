#!/usr/bin/python
#from matplotlib import pyplot as plt
#import matplotlib as mpl
#import numpy as np
from collections import OrderedDict
import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import pdb
import os
import pickle
import sys
from Gopyxl import *
import re
import subprocess

def isNumber(input):
    try:
        a = float(input)
        return True
    except ValueError:
        return False
    except TypeError:
        return False      


def suffix_correction(de_string):
    # Changes differential equation from human-readable format to what Goma expects
    de_string = de_string.replace('-x','1')
    de_string = de_string.replace('-y','2')
    de_string = de_string.replace('-z','3')
    de_string = de_string.replace('-xx','11')
    de_string = de_string.replace('-xy','12')
    de_string = de_string.replace('-xz','13')
    de_string = de_string.replace('-yy','22')
    de_string = de_string.replace('-yz','23')
    de_string = de_string.replace('-zz','33')

    return de_string

if __name__ == '__main__':
    
    flow_ctrl = list()
    
    # Called with no arguments
    if len(sys.argv) == 1:
        # Find the directory of the xlsx file
        cwd_path = os.getcwd()
        this_script_filepath = os.path.realpath(__file__)

        this_script_dir_path, this_script_filename = os.path.split(this_script_filepath)

        # If this script is run in Sublime,
        if cwd_path == this_script_dir_path:
            # Run this xlsx file
            xlsx_filename = '/home/kerry/Research/Projects/THRG_G1/THRG_G1_new.xlsx'
            # xlsx_filename = '/home/kerry/Research/Projects/AMHRG_G12/AMHRG_G12.xlsx'
            

        # Else if it's run in any directory from the temrinal
        else:
            # Run with the first XLSX file it finds
            xlsx_list = [f for f in os.listdir(cwd_path) if f.endswith('.xlsx')]
            xlsx_list = [f for f in xlsx_list if not f.startswith('~$')]
            xlsx_filename = xlsx_list[0]

        # Set paths
        path = os.path.abspath(xlsx_filename)
        path, filename = os.path.split(path)
        drive=''
        print("Option 1")

    # Called with 2 arugments: -s [filename]
    elif "-s" in sys.argv and len(sys.argv) == 3:
        # pdb.set_trace()

        path = os.getcwd()
        filename = sys.argv[2]
        drive=''
        xlsx_filename = path + '/' + filename
        print("Option 2")

    # I don't remember why this option exists
    else:
        filename= sys.argv[1]
        path = os.getcwd()
        xlsx_filename = path + '/' + filename
        drive=''
        print("Option 3")

    file_info = {'drive':drive, 'path':path, 'xlsx':filename}
    
    print("Path: " + path)
    print("Reading: " + xlsx_filename)

    # Load workbook
    wb = openpyxl.load_workbook(xlsx_filename, data_only=True)
    

    # Read worksheets
    Setup_sheet = wb['Setup']
    # Commands_sheet = wb['Commands']
    DiffEQs_sheet = wb['Differential Equations']
    Boundary_conditions_sheet = wb['Boundary Conditions']
    Boundaries_sheet = wb['Boundary Definition']
    Post_processing_sheet = wb['Post-Processing']

    # pdb.set_trace()
    if "-tsurf" in sys.argv:
        surfaces_id_list = [int(s.value) for s in list(Boundaries_sheet.columns)[0][1:]]
        
        for i, rowOfCellObjects in enumerate(Boundaries_sheet['C2':get_column_letter(Boundaries_sheet.max_column) + str(Boundaries_sheet.max_row)]):
            #print i
            segments = [x.value for x in rowOfCellObjects if x.value != None]
            Boundaries_dict[boundaries_prelimid_list[i]] = Boundary(boundaries_prelimid_list[i], boundaries_name_list[i], segments)
     

    # Setup
    # These parts of the Excel file and input file are pretty much just a fixed set of options for
    # any particular command
    # For example: Command = [option]
    # A few of them are set as Data Validation lists in Excel.  Others are text or numbers.
    # Basically we read all these into an OrderedDict from Excel and spit them back out in the input file.
    # FEM File Specifications
    FEM_file_specifications = [s.value.strip() for s in list(Setup_sheet.columns)[0][1:6]]
    FEM_file_specifications_input = [s.value for s in list(Setup_sheet.columns)[1][1:6]]
    FEM_file_dict = OrderedDict(zip(FEM_file_specifications, FEM_file_specifications_input))

    # General Specifications
    General_specifications = [s.value.strip() for s in list(Setup_sheet.columns)[0][8:11]]
    General_specifications_input = [s.value for s in list(Setup_sheet.columns)[1][8:11]]
    General_dict = OrderedDict(zip(General_specifications, General_specifications_input))

    # Time integration specifications
    Time_integration_specifications = [s.value for s in list(Setup_sheet.columns)[0][13:23]]
    Time_integration_specifications_input = [s.value for s in list(Setup_sheet.columns)[1][13:23]]
    Time_integration_dict = OrderedDict(zip(Time_integration_specifications, Time_integration_specifications_input))

    # Solver specifications
    Solver_specifications = [s.value for s in list(Setup_sheet.columns)[3][1:20]]
    Solver_specifications_input = [s.value for s in list(Setup_sheet.columns)[4][1:20]]
    Solver_dict = OrderedDict(zip(Solver_specifications, Solver_specifications_input))

    # Post-Processing
    pp_frequent_specifications = [s.value for s in list(Post_processing_sheet.columns)[0][1:19]]
    pp_frequent_specifications_input = [s.value for s in list(Post_processing_sheet.columns)[2][1:19]]
    pp_porous_specifications = [s.value for s in list(Post_processing_sheet.columns)[4][1:10]]
    pp_porous_specifications_input = [s.value for s in list(Post_processing_sheet.columns)[6][1:10]]

    pp_dict = OrderedDict(zip(pp_frequent_specifications+pp_porous_specifications, \
        pp_frequent_specifications_input+pp_porous_specifications_input))
    
    # pdb.set_trace()       
    
    # # Read boundaries    ==============================================
    Boundaries_dict = dict()
    boundaries_prelimid_list = [int(s.value) for s in list(Boundaries_sheet.columns)[5][1:]]
    boundaries_name_list = [s.value for s in list(Boundaries_sheet.columns)[0][1:]]
    for i in range(len(boundaries_prelimid_list)):
        NS_id = int('1'+str(boundaries_prelimid_list[i]))
        SS_id = int('2'+str(boundaries_prelimid_list[i]))

        Boundaries_dict[NS_id] = Boundary(NS_id, boundaries_name_list[i], 'NS')
        Boundaries_dict[SS_id] = Boundary(SS_id, boundaries_name_list[i], 'SS')
    
    # Read Problem Description
    # Read lines that say Material Name
    inactive_col_rd = [(s.value) for s in list(DiffEQs_sheet.columns)[0][0:]] # Col A
    mat_name_col_rd = [(s.value) for s in list(DiffEQs_sheet.columns)[1][0:]] # Col B
    options_col_rd = [(s.value) for s in list(DiffEQs_sheet.columns)[2][0:]]  # Col C
    diffeqs_col_rd = [(s.value) for s in list(DiffEQs_sheet.columns)[4][0:]]  # Col E

    mat_name_rows  = [i for i,x in enumerate(mat_name_col_rd) if x == 'Material name']
    inactive_rows = [i for i,x in enumerate(inactive_col_rd) if x == 1]

    # active_mat_name_rows = [row for row in mat_name_rows if row not in inactive_rows]
    # # mat_name_lines = 

    MatField_list = list()

    # Read which Materials are active
    for i, mat_row in enumerate(mat_name_rows):
        if inactive_col_rd[mat_row]==1:
            matfield_active = False
        else:
            matfield_active = True

        print ('Reading Material Name in row {}'.format(mat_row))
        read_opt_row_begin = mat_name_rows[i]+1

        if i == len(mat_name_rows)-1:
            read_opt_row_end = DiffEQs_sheet.max_row

        else:
            read_opt_row_end = mat_name_rows[i+1]

            mat_options = options_col_rd[read_opt_row_begin:read_opt_row_end]

        material_name = options_col_rd[mat_row] # The material name
        opts_dict = OrderedDict() # Fill this ordered OrderedDict 
        diffeq_list = list()

        # Read each of the rows between the ones with active Materials
        for opt_row in range(read_opt_row_begin, read_opt_row_end):
            # If we encounter an inactive material row, break out of this loop

            if mat_name_col_rd[opt_row] is not None:
                opts_dict[mat_name_col_rd[opt_row]] = options_col_rd[opt_row]

            if diffeqs_col_rd[opt_row] is not None:
                print ('Reading EQs in row {0}'.format(opt_row))
                # DiffEQ command name
                diffeq_name = suffix_correction(diffeqs_col_rd[opt_row])

                # DiffEQ active
                diffeq_active_rd = list(DiffEQs_sheet.rows)[opt_row][3].value

                if diffeq_active_rd == 1:
                    diffeq_active = False
                else:
                    diffeq_active = True

                # Array of strings: 0. Variable name  1. Galerkin wt fn  2. Interpolation fn
                diffeq_var_fns = [s.value for s in list(DiffEQs_sheet.rows)[opt_row][5:8]] 
                # Array of floats for various weights
                diffeq_weights = [s.value for s in list(DiffEQs_sheet.rows)[opt_row][8:17]]

                diffeq_list.append(DiffEq(diffeq_name, diffeq_active, diffeq_var_fns[0], diffeq_var_fns[1], diffeq_var_fns[2], diffeq_weights))

        MatField_list.append(MaterialField(material_name, opts_dict, matfield_active, diffeq_list))


    # Boundary Conditions   =============================================
    # bc_id_list = [int(s.value) for s in list(Boundary_conditions_sheet.columns)[column_index_from_string('A')-1][1:] if s.value is not None]
    # bc_list =  [s.value for s in list(Boundary_conditions_sheet.columns)[column_index_from_string('C')-1][1:] if s.value is not None]
    # bc_active_list =  [s.value for s in list(Boundary_conditions_sheet.columns)[column_index_from_string('B')-1][1:]]

    bc_id_list = [int(s.value) if s.value is not None else -1 for s in list(Boundary_conditions_sheet.columns)[column_index_from_string('A')-1][1:]]
    bc_list =  [s.value for s in list(Boundary_conditions_sheet.columns)[column_index_from_string('C')-1][1:]]
    bc_active_list =  [s.value for s in list(Boundary_conditions_sheet.columns)[column_index_from_string('B')-1][1:]]
    
    custom_boundary_id = Boundary_conditions_sheet['I1'].value

    # raw = [s.value for s in list(Boundary_conditions_sheet.columns)[column_index_from_string('H')-1][1:]]
    # pdb.set_trace()

    # If no custom boundary IDs
    if custom_boundary_id == 'FALSE':
        print('Reading autogenerated boundary IDs')
        # Read boundary IDs from column H
        bc_boundary_id_list =  [int(s.value) if bool(s.value) else -1 for s in list(Boundary_conditions_sheet.columns)\
            [column_index_from_string('H')-1][1:]]
    else:
        print('Reading custom boundary IDs')
        # Read Boundary IDs from from column I
        bc_boundary_id_list =  [int(s.value) if bool(s.value) else -1 for s in list(Boundary_conditions_sheet.columns)\
            [column_index_from_string('I')-1][1:]]

    bc_int_data_str = [str(s.value).replace(" ", "").split(',') for s in list(Boundary_conditions_sheet.columns)\
        [column_index_from_string('J')-1][1:]]

    #pdb.set_trace()
    active_dict = {None:True, 1:False} #Basically converts 0 => True and 1 => False.  Technically shows inactivity...

    # Read arguments for each BC
    bc_argument_list = [str(s.value) for s in list(Boundary_conditions_sheet.columns)\
        [column_index_from_string('J')-1][1:]]

    # Read expected arguments for each BC (not required, but may come in handy later)
    bc_arg_expect_list = [str(s.value) for s in list(Boundary_conditions_sheet.columns)\
        [column_index_from_string('K')-1][1:]]
    
    BC_dict = dict() # Dictionary of BCs.  Keys are BC id.  Items are BC.

    # print len(bc_id_list)
    # print len(bc_active_list)
    # print len(bc_boundary_id_list)
    # print len(bc_int_data_list)
    # print len(bc_real_data_list)
    # print bc_boundary_id_list

    # print("i:")

    # Make sure number of BCs (aka number of BC IDs) match the number of mesh Boundary IDs
    if len(bc_id_list)!= len(bc_boundary_id_list):
        print("*** Check that all BCs are assigned a boundary! ***")

    # For each Boundary ID (aka row where there is a BC),
    # populate BC_dict with BoundaryCondition() objects as values and the BC ID as the key
    for i in range(len(bc_id_list)):

        # print bc_boundary_id_list[i] 
        # print i 
        # pdb.set_trace()
        skip_condition = bc_list[i] is 'CATEGORY' or bc_boundary_id_list[i] < 0

        if skip_condition:
            # pdb.set_trace()
            print('Skipping: BC ' + str(bc_id_list[i]) + ' ' + str(bc_list[i]))
            pass
        else:
            BC_dict[bc_id_list[i]] = BoundaryCondition(bc_id_list[i], active_dict[bc_active_list[i]], Boundaries_dict[bc_boundary_id_list[i]], \
            bc_list[i], bc_argument_list[i], bc_arg_expect_list[i])

        #print "bc(" + str(bc_id_list[i]) + ', ' + bc_list[i].lower() + ", (" + \
        #str(bc_boundary_id_list[i]) + ", (" + str(len(bc_int_data_list[i])) + ", " + \
        #str(bc_int_data_list[i]).strip("[]") + "), (" + str(len(bc_real_data_list[i])) + ", ", str(bc_real_data_list[i]).strip("[]") + "))"
    
    # # Custom Lines
    # Custom_lines = [s.value for s in list(Custom_lines_sheet.columns)[0][0:]]

    # Dump all this info into a Gopyxl() object
    gopyxl = Gopyxl(file_info, FEM_file_dict, General_dict, Time_integration_dict, \
        Solver_dict, MatField_list, Boundaries_dict, BC_dict, pp_dict)

    # Print the input file
    gopyxl.print_goma_input()
    print(path)
    os.chdir(path)

    # Write and save the input file
    gopyxl.write_goma_input()
    gopyxl.save()

    # If this is being run locally, osync the directory where this is being executed.
    if os.path.isfile('/home/kerry/localcheck'):
        print('Running locally.  Osync this automatically.')
        subprocess.call(['/bin/bash', '-i', '-c', 'othis'])
        # os.system('othis')

# for key, value in pp_dict.items():
#     print key, value

# pdb.set_trace()







